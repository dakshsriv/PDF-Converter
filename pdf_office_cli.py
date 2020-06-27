#!/usr/bin/python3

import argparse
import fpdf
import magic
from docx2pdf import convert
import xtopdf
from pdfrw import PdfReader, PdfWriter
from openpyxl import load_workbook
import os
import reportlab

def findFileType(filename):
    i = int()
    for x in range(-1, -1* len(filename), -1):
        i = i + 1
        if filename[x] != ".":
            break
    i = i - 1
    i = -1 * i
    filetype = filename[i:]
    return filetype

def createPDF(filename):
    filetype = findFileType(filename)
    print(filetype)
    #print(os.getcwd())
    i = 0
    for x in range(-1, -1* len(filename), -1):
        print(i)
        i = i + 1
        if filename[x] != "/":
            break
    i = i - 1
    i = -1 * i
    print(filename[0: int(i)])
    os.chdir(filename[0: int(i)])
    print(os.getcwd())
    if filetype == "docx":
        convert(filename[i+1:], filename[0: int(i)] + filename[i+1:-4] + "pdf")
    elif filetype == "xlsx":
        """
        workbook = load_workbook('fruits2.xlsx', guess_types=True, data_only=True)
        worksheet = workbook.active

        pw = PdfWriter('fruits2.pdf')
        pw.setFont('Courier', 12)
        pw.setHeader('XLSXtoPDF.py - convert XLSX data to PDF')
        pw.setFooter('Generated using openpyxl and xtopdf')

        ws_range = worksheet.iter_rows('A1:H13')
        for row in ws_range:
            s = ''
            for cell in row:
                if cell.value is None:
                    s += ' ' * 11
                else:
                    s += str(cell.value).rjust(10) + ' '
            pw.writeLine(s)
        pw.savePage()
        pw.close()"""
        pass



if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "-c",
        "--create",
        type=str,
        default="",
        help="Enter the command followed by the file you want to convert to pdf. Docx, Xlsx, and Txt only. Example: Cli_version.py -create home/username/Documents/Notes/Note.docx",
    )
    parser.add_argument(
        "-e",
        "--multiedit",
        type=str,
        default="",
        help="Enter multirhythm settings. Example: 4,5,60 for Ratio of 4 to 5 at 60 bpm.",
    )
    parser.add_argument("-p", "--play", help="Play?", action="store_true")
    parser.add_argument("-play", "--multiplay", help="Play?", action="store_true")
    args = parser.parse_args()
    if args.create:
        createPDF(args.create)